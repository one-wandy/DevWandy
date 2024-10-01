import os
from . import models
from django.views.generic import *
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.shortcuts import redirect
from Customer.mixing import Options
from datetime import datetime
from django.template.loader import get_template
from django.http import HttpResponseRedirect, HttpResponse
from fpdf import FPDF
from django.http import JsonResponse
from twilio.rest import Client  
from Customer import models
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
# App Account
from oauth2client.service_account import ServiceAccountCredentials
from allauth.socialaccount.models import SocialAccount
from google.oauth2.credentials import Credentials

import shutil

from google.oauth2 import service_account


from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# Suponiendo que ya tienes un objeto 'credentials' válido

class Maps(TemplateView, Options):
      template_name = "maps/information.html"

              
      def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['setting'] = self.Setting()
        return context
    
    
class DataCredit(TemplateView, Options):
      template_name = "maps/data-credit.html"

              
      def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['setting'] = self.Setting()
        return context
    
      
      def get(self, request, *args, **kwargs):
            nombre_archivo = "Info.docx"
            ruta_carpeta =  os.getcwd() + "/Clientes"

            # self.Send_WhatsApp_Message()
            load_excel = 'Customer/Archivos/DataCredit.xlsx'

            return  self.LoadExcel(load_excel)
            # # Ejemplo de uso
            nombre_contacto = "Juan Perez"
            telefono_contacto = "+1234567890"
            try:
            # Llamada a la función para agregar un contacto
                  resultado = self.add_contacts_to_google_contact()

            # Si no se lanza ninguna excepción, se asume que la creación del contacto fue exitosa
                  print("El contacto se agregó correctamente:", resultado)
            except Exception as e:
            # Si se produce una excepción, se captura y se imprime el mensaje de error
                  print("Error al agregar el contacto:", e)

            # self.agregar_contacto(nombre_contacto, telefono_contacto)

            return super().get(request, *args, **kwargs)
      

      def Send_WhatsApp_Message(self):
            account_sid = 'AC32b5e94ce632aabd0a278a56e16bd44a'
            auth_token = 'a61c8b622e93ada5958d69dacef7c461'
            client = Client(account_sid, auth_token)

            Msm = "Recordario de Grupo Fycas \n \n Buenos días estimado(a), este es un recordatorio de su saldo pendiente \n \n Nombre: Xavier Torrero \n Cedula: 302-234433-0 \n Numero: (809)299-8306 \n \n Nos ponemos en contacto con usted para dar seguimiento al pago Numero dos (2) de veinte (20) cuotas. \n\n La fecha de vencimiento de su ultimo pago fue el 30/3/2024, a dia de hoy 6/4/2024 no hemos recibido el pago correspondiente. \n \n Para realizar el pago de la cuota pendiente mas mora actualmente es de; RD$5,500.00, puede llamar o dejar un mensaje via WhatsApp al +1 (809)870-7852. \n Atentamente: \n (Grupo Fycas)"
            message = client.messages.create(
                  body= Msm,
                  from_= '+13344384583',
                  to= '+18295577196' )
            
            # msg = client.messages.create(
            #       from_='whatsapp:+18295577196',
            #       body='Mensaje enviado por Wandy Olivares',
            #       to='whatsapp:+18295577196')
            return True



      def LoadExcel(self, existing_excel_path):
            # Obtener la fecha y hora actual
            date = datetime.now()
            formatted_date = date.strftime("%Y-%m-%d_%H-%M-%S")

            # Crear una copia del archivo existente
            new_excel_path = f"Copia{formatted_date}.xlsx"
            shutil.copy(existing_excel_path, new_excel_path)

            # Cargar la copia del archivo Excel existente.
            workbook = load_workbook(new_excel_path)

            # Seleccionar la primera hoja del libro de trabajo.
            worksheet = workbook.active

            # Definir la fila inicial para escribir los datos (fila 7 en este caso)
            start_row = 7

            # Obtener los datos de Django.
            data = models.Customer.objects.filter(is_active = True, 
                                    customer_verify = True ).order_by('-id')
            print(data)

            # Iterar sobre los datos y escribirlos en la copia del archivo Excel a partir de la fila inicial.
            for index, row_data in enumerate(data, start=start_row):
                  worksheet.cell(row=index, column=1).value = row_data.type_input
                  worksheet.cell(row=index, column=2).value = row_data.name
                  worksheet.cell(row=index, column=3).value = row_data.last_name
                  worksheet.cell(row=index, column=4).value = row_data.dni
                  worksheet.cell(row=index, column=5).value = row_data.sexo
                  worksheet.cell(row=index, column=6).value = row_data.estado_civil
                  worksheet.cell(row=index, column=7).value = row_data.ocupacion
                  worksheet.cell(row=index, column=8).value = row_data.code_customer
                  worksheet.cell(row=index, column=9).value = row_data.nacimiento
                  worksheet.cell(row=index, column=10).value = row_data.nacionalidad
                  worksheet.cell(row=index, column=11).value = row_data.direccion
                  worksheet.cell(row=index, column=12).value = row_data.sector
                  worksheet.cell(row=index, column=13).value = row_data.calle_numero
                  worksheet.cell(row=index, column=14).value = row_data.municipio
                  worksheet.cell(row=index, column=15).value = row_data.ciudad
                  worksheet.cell(row=index, column=16).value = row_data.provincia
                  worksheet.cell(row=index, column=17).value = row_data.pais
                  worksheet.cell(row=index, column=18).value = row_data.dir_referencia
                  worksheet.cell(row=index, column=19).value = row_data.number
                  worksheet.cell(row=index, column=20).value = row_data.phone
                  worksheet.cell(row=index, column=21).value = row_data.empresa_trabaja
                  worksheet.cell(row=index, column=22).value = row_data.cargo
                  worksheet.cell(row=index, column=23).value = row_data.direccion_trabajo
                  worksheet.cell(row=index, column=24).value = row_data.sector
                  worksheet.cell(row=index, column=25).value = row_data.calle_numero_trabajo
                  worksheet.cell(row=index, column=26).value = row_data.municipio_trabaja
                  worksheet.cell(row=index, column=27).value = row_data.ciudad_trabaja
                  worksheet.cell(row=index, column=28).value = row_data.provincia_trabajo
                  worksheet.cell(row=index, column=29).value = row_data.pais_trabajo
                  worksheet.cell(row=index, column=30).value = row_data.dir_referencia_trabajo
                  worksheet.cell(row=index, column=31).value = row_data.salario_m
                  worksheet.cell(row=index, column=32).value = row_data.moneda
                  worksheet.cell(row=index, column=33).value = row_data.relacion_tipo
                  worksheet.cell(row=index, column=34).value = row_data.fecha_apertura
                  worksheet.cell(row=index, column=35).value = row_data.fecha_vencimiento
                  worksheet.cell(row=index, column=36).value = row_data.fecha_ultimo_pago
                  worksheet.cell(row=index, column=37).value = row_data.numeoro_cuenta
                  worksheet.cell(row=index, column=38).value = row_data.estatus
                  worksheet.cell(row=index, column=39).value = row_data.tipo_prestamo
                  worksheet.cell(row=index, column=40).value = row_data.moneda_prestamo
                  worksheet.cell(row=index, column=41).value = row_data.credito_aprovado
                  worksheet.cell(row=index, column=42).value = row_data.balance_corte
                  worksheet.cell(row=index, column=43).value = row_data.monto_adeudado
                  worksheet.cell(row=index, column=44).value = row_data.pago_mandatorio_cuota
                  worksheet.cell(row=index, column=45).value = row_data.monto_ultimo_pago
                  worksheet.cell(row=index, column=46).value = row_data.total_atraso
                  worksheet.cell(row=index, column=47).value = row_data.tasa_interes
                  worksheet.cell(row=index, column=48).value = row_data.forma_pago
                  worksheet.cell(row=index, column=49).value = row_data.cantidad_cuota
                  worksheet.cell(row=index, column=50).value = row_data.atraso1_30
                  worksheet.cell(row=index, column=51).value = row_data.atraso31_60
                  worksheet.cell(row=index, column=52).value = row_data.atraso61_90
                  worksheet.cell(row=index, column=53).value = row_data.atraso91_120
                  worksheet.cell(row=index, column=54).value = row_data.atraso121_150
                  worksheet.cell(row=index, column=55).value = row_data.atraso151_180
                  worksheet.cell(row=index, column=56).value = row_data.atraso181_o_mas


            # Guardar los cambios en la copia del archivo Excel.
            workbook.save(new_excel_path)

            # Preparar la respuesta HTTP con la copia del archivo Excel para descargar.
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=DataCredit{formatted_date}.xlsx'
            with open(new_excel_path, 'rb') as file:
                  response.write(file.read())

            return response

      def Excel(self):
            # Consultar el modelo de Django para obtener los datos.
            data = models.Customer.objects.filter(is_active = True, 
                                    customer_verify = True ).order_by('-id')
            date = datetime.now()
            
            # Crear un nuevo archivo Excel.
            workbook = Workbook()

            # Escribir los datos en el archivo Excel.
            for nombre_hoja in workbook.sheetnames:
                  worksheet = workbook[nombre_hoja]
                  bold_font = Font(bold=True)
                  
                  # Ajustar el ancho de las columnas
                  columnas = ["A", "B", "C", "D", "E", "F", "G"]
                  for columna in columnas:
                        worksheet.column_dimensions[columna].width = 20
                  worksheet.column_dimensions["A"].width = 25
                  worksheet.column_dimensions["A"].font = bold_font
                  
                  worksheet.column_dimensions["B"].width = 30
                  worksheet.column_dimensions["C"].width = 20
                  worksheet.column_dimensions["D"].width = 20
                  worksheet.column_dimensions["E"].width = 10
                  worksheet.column_dimensions["F"].width = 20
                  worksheet.column_dimensions["G"].width = 20
                  # Aplicar formato de fuente a la fila 3


                  worksheet.append(["TIPO DE ENTIDAD", "NOMBRE DEL CLIENTE", "APELLIDOS", "CEDULA O RNC", "SEXO", "ESTADO CIVIL", "OCUPACION", "CODIGO DE CLIENTE", "FECHA DE NACIMIENTO", "NACIONALIDAD", "DIRECCION", "SECTOR", "CALLE/NUMERO", "MUNICIPIO", "CIUDAD", "PROVINCIA", "PAIS", "DIR_Referencia", "TELEFONO", "TELEFONO", 'EMPRESA DONDE TRABAJA', "CARGO", "DIRECCION", "SECTOR", "CALLE/NUMERO", "MUNICIPIO", "CIUDAD", "PROVINCIA", "PAIS", "Dir_Referencia",  "EMPRESA DONDE TRABAJA, 'CARGO", 'DIRECCION', 'SECTOR', 'CALLE/NUMERO', 'MUNICIPIO', "CIUDAD", "PROVINCIA", "Pais", "Dir_Referencia", "SALARIO MENSUAL", "MONEDA SALARIO", "RELACION TIPO", "FECHA APERTURA", "FECHA VENCIMIENTO", "FECHA ULTIMO PAGO", "NUMERO CUENTA", "ESTATUS", "TIPO DE PRESTAMO", "MONEDA", "CREDITO APROBADO", "BALANCE AL CORTE MONTO ADEUDADO", "PAGO MANDATORIO O CUOTA", "MONTO ULTIMO PAGO", "TOTAL DE ATRASO", "TASA DE INTERES", "FORMA DE PAGO", "CANTIDAD DE CUOTA", "ATRASO 1 A 30 DIAS", "ATRASO 31 A 60 DIAS", "ATRASO 61 A 90 DIAS", "ATRASO 91 A 120 DIAS", "ATRASO 121 A 150 DIAS", "ATRASO 151 A 180 DIAS", "ATRASO 181 Dias o MAS "  ])

                  # Insertar una nueva fila en la fila 1
                  worksheet.insert_rows(1)
                  # Combinar las celdas de la fila 1
                  worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
         
                  # Insertar una nueva fila en la fila 2
                  worksheet.insert_rows(2)
                  worksheet.insert_rows(3)
                  worksheet.insert_rows(4)
                  worksheet.insert_rows(5)
                  # Combinar las celdas de la fila 2
                  worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)

                  worksheet.cell(row=1, column=1).value = "GRUPO FYCAS, SRL"
                  worksheet.cell(row=1, column=1).font = Font(name="Arial", size=18, bold=True, color="000000")

                  worksheet.cell(row=2, column=1).value = "AF000024759"
                  worksheet.cell(row=2, column=1).font = Font(name="Arial", size=14)
                            
                  for cell in worksheet[1]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[2]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[3]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[4]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[5]:
                     cell.fill = PatternFill(start_color="0080ff", end_color="0080ff", fill_type="solid")
   
                  
                  worksheet.cell(row=5, column=1).value = "Datos Personales"
                  worksheet.cell(row=5, column=1).font =  Font(name="Cambria", size=16, bold=True, color="FFFFFF")
                  
         
                  for col in range(1, 68):  # Iterate from column B to G (1-based indexing)
                        cell = worksheet.cell(row=6, column=col)
                        cell.font = Font(name="Calibri", size=14,  bold=True)
                        
                  for row in data:
                        worksheet.append([row.type_input, row.name, row.last_name, row.dni, row.sexo, row.estado_civil, row.ocupacion, row.code_customer, row.nacimiento, row.nacionalidad,  row.direccion, row.sector, row.calle_numero, row.municipio,  row.ciudad, row.provincia, row.pais, row.dir_referencia, row.number, row.phone, row.empresa_trabaja, row.cargo, row.direccion_trabajo, row.sector, row.calle_numero_trabajo, row.municipio_trabaja, row.ciudad_trabaja, row.provincia_trabajo, row.pais_trabajo, row.dir_referencia_trabajo, row.salario_m, row.moneda, row.relacion_tipo, row.fecha_apertura, row.fecha_vencimiento, row.fecha_ultimo_pago, row.numeoro_cuenta, row.estatus, row.tipo_prestamo, row.moneda_prestamo, row.credito_aprovado, row.balance_corte, row.monto_adeudado, row.pago_mandatorio_cuota, row.monto_ultimo_pago, row.total_atraso, row.tasa_interes, row.forma_pago, row.cantidad_cuota, row.atraso1_30, row.atraso31_60, row.atraso61_90, row.atraso91_120, row.atraso121_150, row.atraso151_180, row.atraso181_o_mas])
                        
                        

            # Enviar el archivo Excel al usuario como respuesta HTTP.
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Grupo Fycas SRL {date}.xlsx'
            workbook.save(response)
            return response
      
            # views.py
      
      # Define los ámbitos de la API de Google Contacts que necesitas

      def add_contacts_to_google(self):
            SCOPES = ['https://www.googleapis.com/auth/contacts']
            # Carga las credenciales almacenadas (si existen)
            creds = None
            if os.path.exists('credentials.json'):
                  creds = Credentials.from_authorized_user_file('credentials.json')

            # Si no hay credenciales disponibles o están vencidas, solicita al usuario que inicie sesión
            if not creds or not creds.valid:
                  if creds and creds.expired and creds.refresh_token:
                        creds.refresh(Request())
                  else:
                        flow = InstalledAppFlow.from_client_secrets_file('Customer/Credentials-Apis/credentials.json', SCOPES, redirect_uri='http://localhost:56083')
                        creds = flow.run_local_server(port=63880)

                  # Guarda las credenciales para futuros usos
                  with open('token.json', 'w') as token:
                        token.write(creds.to_json())

            # Construye el servicio de la API de Google Contacts
            service = build('people', 'v1', credentials=creds)

            # Obtiene los contactos desde tu modelo Django
            # customers = Customer.objects.all()

            # Agrega cada contacto a Google Contacts
            # for customer in customers:
            contact = {
                  'names': [
                  {
                        'givenName': 'Federico',
                        'familyName': 'Almanzar'
                  }
                  ],
                  'emailAddresses': [
                  {
                        'value': 'federico@gmail.com'
                  }
                  ]
                              # Puedes agregar más campos del contacto aquí según sea necesario
                  }

                  # Llama al método 'createContact' de la API de Google Contacts para agregar el contacto
            contact = service.people().createContact(body=contact).execute()
            print('Contacto agregado:', contact)
            return True
# Llama a la función para agregar los contactos a Google Contacts
# add_contacts_to_google()


      def add_contacts_to_google_contact(self):
            credenciales = service_account.Credentials.from_service_account_file('Customer/Credentials-Apis/credentials.json')

            persona =  {
                  'names': [{
                        'givenName': 'Wandys',
                  }],
                  'emailAddresses': [{
                        'value': 'wandy.one@gmail.com',
                        'type': 'home',
                  }],
                  'phoneNumbers': [{
                        'value': '809-001-0011',
                        'type': 'home',
                  }],
            }
            cliente = build('people', 'v1', credentials=credenciales)
            cliente.people().createContact(body=persona).execute()
            # contact = service.people().createContact(body=new_contact).execute()
            return True




      def obtener_credenciales(self):
            SCOPES = ['https://www.googleapis.com/auth/contacts']
            
            # Intenta cargar las credenciales desde el archivo token.json
            creds = None
            if os.path.exists('credentials.json'):
                  creds = Credentials.from_authorized_user_file('credentials.json')

            # Si no hay credenciales disponibles o no son válidas, inicia el flujo de autorización
            if not creds or not creds.valid:
                  if creds and creds.expired and creds.refresh_token:
                        creds.refresh(Request())
                  else:
                        # Si no hay credenciales disponibles, inicia el flujo de autorización
                        flow = InstalledAppFlow.from_client_secrets_file('Customer/Credentials-Apis/credentials.json', SCOPES, redirect_uri='http://localhost:56083')
                        
                        creds = flow.run_local_server(port=56083)

                  # Guarda las credenciales en credentials.json para futuros usos
                  with open('credentials.json', 'w') as token:
                        token.write(creds.to_json())

            return creds

      def agregar_contacto(self, nombre, telefono):
            creds = self.obtener_credenciales()
            # Crea un objeto de servicio para interactuar con la API de Google Contacts
            service = build('people', 'v1', credentials=creds)
            # Define el contacto
            contacto = {
                  "names": [
                        {
                        "givenName": nombre
                        }
                  ],
                  "phoneNumbers": [
                        {
                        "value": telefono,
                        "type": "mobile"
                        }
                  ]
            }
            # Agrega el contacto
            service.people().createContact(body=contacto).execute()
            print("Contacto agregado exitosamente.")
